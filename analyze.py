import openpyxl
import json
import sys

def extract_excel_info(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        info = {'filename': filepath, 'sheets': {}}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= 5: # Get up to 5 rows
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])
            info['sheets'][sheet_name] = rows
        return info
    except Exception as e:
        return {'filename': filepath, 'error': str(e)}

file1 = "Copia de IAD 158 Control Inspecciones 2.xlsx"
file2 = "Copia de IAD_157,158-JOB_LOG_00.xlsx"

print("================ FILE 1 ================")
info1 = extract_excel_info(file1)
print(json.dumps(info1, indent=2))

print("\n================ FILE 2 ================")
info2 = extract_excel_info(file2)
print(json.dumps(info2, indent=2))
