import os
import openpyxl
from supabase import create_client, Client

# =========================================================================
# SCRIPT DE MIGRACIÓN: EXCEL A SUPABASE
# Para ejecutar este script necesitas instalar:
# pip install openpyxl supabase
# =========================================================================

# Configuración de Supabase (Reemplaza con tus credenciales)
url: str = os.environ.get("SUPABASE_URL", "https://ssfwmtjftfwlsfvdvzex.supabase.co")
key: str = os.environ.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNzZndtdGpmdGZ3bHNmdmR2emV4Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODYyNDE3MTMsImV4cCI6MjEwMTgxNzcxM30.yFF6VZC7UCZTpi9FgCPHw-C2hVtRzxzSObW8GRpQ8lQ")
supabase: Client = create_client(url, key)

file1 = "Copia de IAD 158 Control Inspecciones 2.xlsx"
file2 = "Copia de IAD_157,158-JOB_LOG_00.xlsx"

def migrate_inspections():
    print(f"Migrando Inspecciones desde {file1}...")
    try:
        wb = openpyxl.load_workbook(file1, data_only=True)
        if "Registro de Inspecciones" in wb.sheetnames:
            sheet = wb["Registro de Inspecciones"]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0: continue # Saltar encabezados (asumiendo que row 0 es header)
                
                # Extraer valores, asumiendo estructura:
                # [ID, Tipo, Area, Nivel, Elemento, Fecha Programada, Fecha Ejecutada, Status, Comentarios, Inspector]
                # Nota: Ajusta los índices (row[X]) según las columnas reales del Excel.
                if row[0] is None: continue # Fila vacía

                data = {
                    "element": str(row[4]) if len(row) > 4 and row[4] else "N/A",
                    "status": str(row[7]) if len(row) > 7 and row[7] else "Pendiente",
                    "comments": str(row[8]) if len(row) > 8 and row[8] else ""
                }
                # Insertar en Supabase (Omitiendo FKs como area_id por ahora para el ejemplo base)
                response = supabase.table("inspections").insert(data).execute()
                print(f"Insertado: {data['element']}")
    except Exception as e:
        print(f"Error migrando inspecciones: {e}")

def migrate_materials():
    print(f"Migrando Materiales desde {file2}...")
    try:
        wb = openpyxl.load_workbook(file2, data_only=True)
        if "Material LOG (158)" in wb.sheetnames:
            sheet = wb["Material LOG (158)"]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i < 4: continue # Saltar filas vacías y encabezados de este archivo
                if row[0] is None: continue # Fila vacía
                
                # Encabezados: [SUBMITTAL, DESCRIPTION, Model, LEAD TIME, REQUIRED ONSITE, REQUIRED RELEASE, ACTUAL RELEASE, EXPECTED DELIVERY, ISSUES]
                data = {
                    "submittal_number": str(row[0]),
                    "description": str(row[1]) if len(row)>1 and row[1] else "",
                    "model": str(row[2]) if len(row)>2 and row[2] else "",
                    "issues_comments": str(row[8]) if len(row)>8 and row[8] else ""
                }
                response = supabase.table("materials").insert(data).execute()
                print(f"Insertado Material: {data['submittal_number']}")
    except Exception as e:
        print(f"Error migrando materiales: {e}")

if __name__ == "__main__":
    print("Iniciando migración a Supabase...")
    # Asegúrate de haber creado las tablas en Supabase con el script SQL antes de ejecutar esto.
    migrate_inspections()
    migrate_materials()
    print("Migración finalizada.")
