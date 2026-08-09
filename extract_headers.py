import openpyxl
import json
import sys

def extract_headers(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        info = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = []
            sample_data = []
            
            # Find the first row that looks like a header (has multiple string values)
            header_row_idx = -1
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                str_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                if str_count > 3: # Assuming a header row has at least 4 columns
                    header_row_idx = i
                    headers = [str(cell).strip().replace('\n', ' ') if cell else f"Col_{j}" for j, cell in enumerate(row)]
                    break
            
            # Get next 2 rows of data
            if header_row_idx != -1:
                for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx+2, values_only=True)):
                    if i >= 2: break
                    sample_data.append([str(cell) if cell is not None else "" for cell in row])
            
            info[sheet_name] = {
                'headers': headers,
                'sample': sample_data
            }
        return info
    except Exception as e:
        return {'error': str(e)}

file1 = "Copia de IAD 158 Control Inspecciones 2.xlsx"
file2 = "Copia de IAD_157,158-JOB_LOG_00.xlsx"

out = {
    'file1': extract_headers(file1),
    'file2': extract_headers(file2)
}

with open("headers_analysis.json", "w", encoding="utf-8") as f:
    json.dump(out, f, indent=2, ensure_ascii=False)
    print("Analysis saved to headers_analysis.json")
